import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
class Xlsx:
    def __init__(self, filePath):
        self.path = filePath
    
    def writeWorkbook(self, source):
        self.wb = Workbook()
        ws = self.wb.active
        for each in source:
            ws.append(each)
        self.wb.save(self.path)
            
    def writeComment(self, **comments):
        for each in comments:
            self.ws["{}".format(each)].comment = Comment('{}'.format(comments[each]), 'Ziplux')
        self.wb.save(self.path)
            

    def readWorkbook(self):
        self.wb = openpyxl.load_workbook(self.path)
        memory.msm.info(tagName='readWorkbook(filePath) :', tag=self.path)
    
        return self.wb.sheetnames
    
    def readWorksheet(self, sheetname):
        memory.msm.info(tagName='sheetname :', tag=sheetname)
        ws = self.wb[sheetname]
        df = pd.DataFrame(ws.values)
        return df